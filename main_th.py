from PyQt5.QtCore import QThread, pyqtSignal
from openpyxl import load_workbook

class WorkThread(QThread):
    endSignal = pyqtSignal()    # 自定义结束线程信号
    progressSignal = pyqtSignal(int)    # 自定义进度条信号

    def __init__(self, content, path, startDate, endDate):
        super(WorkThread, self).__init__()
        self.content = content
        self.path = path
        self.startDt = startDate
        self.endDt = endDate

    # 重写run方法
    def run(self):
        wb = load_workbook(self.path)   # 打开模板Excel文件
        ws = wb['QQ群服务登记']

        for cell in ws[f'b5:b{ws.max_row}']:
            try:    # 报错跳过本次循环
                idx = self.getIndex(cell[0].value, self.content)
            except:
                continue
            s = self.content[idx[0]:idx[1]]
            currentDate = self.startDt
            endDate = self.endDt
            while currentDate <= endDate:
                row = cell[0].row
                column = currentDate.dayOfYear() + 4
                # 计算次数
                ws.cell(row, column).value = s.count(currentDate.toString('yyyy-MM-dd'))
                currentDate = currentDate.addDays(1)
            self.progressSignal.emit(len(ws[f'b5:b{ws.max_row}']))  # 发送进度条信号

        wb.save(self.path)
        self.endSignal.emit()   # 发送结束信号

    # 获得内容范围索引
    def getIndex(self, name, newStr):
        start = newStr.index(f'消息对象:{name}') + len(f'消息对象:{name}') + 67
        end = newStr.index('================================================================', start)
        return (start, end)
